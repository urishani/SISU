"""Read colored Excel header columns and append book rows into those fields."""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any
import re
import shutil

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Header names in the master file → keys produced by the crawler.
HEADER_TO_FIELD: dict[str, str] = {
    "supplier": "supplier",
    "catalog for mom": "catalog_mom",
    "cat number short": "cat_number",
    "publisher": "publisher",
    "title (phonetics)": "title_phonetic",
    "author (english)": "author_en",
    "title in english": "title_en",
    "title in hebrew": "title_he",
    "author (hebrew)": "author_he",
    "category #1": "category",
    "category #2": "category2",
    "category #3": "category3",
    "upc": "upc",
    "danacode": "danacode",
    "isbn": "isbn",
    "isbn -old": "isbn_old",
    "item type": "item_type",
    "language": "language",
    "translated": "translated",
    "copyright year": "year",
    "number of pages": "pages",
    "cover type: s;h;bb": "cover_type",
    "cover type": "cover_type",
    "spine color": "spine_color",
    "weight (labs)": "weight_lb",
    "weight (oz)": "weight_oz",
    "weight (kg)": "weight_kg",
    "size- hight (inc)": "height_in",
    "size-width (inc)": "width_in",
    "size- thickness (inc)": "thickness_in",
    "size- hight (cm)": "height_cm",
    "size- height (cm)": "height_cm",
    "size-width (cm)": "width_cm",
    "size- width (cm)": "width_cm",
    "size- thickness (cm)": "thickness_cm",
    "israeli price (shekel)": "price_ils",
    "description (english)": "description_en",
    "description (hebrew)": "description_he",
    "cover image url": "cover_image_url",
    "cover page image url": "cover_image_url",
    "cover page url": "cover_image_url",
    "back image url": "back_image_url",
    "back page image url": "back_image_url",
    "back page url": "back_image_url",
    "scanner id": "scanner_id",
    "scannerid": "scanner_id",
    "comments": "comments",
    "keywords": "keywords",
}

ORANGE_FILL = PatternFill(fill_type="solid", fgColor="FFFFC000")
REQUIRED_COLORED_HEADERS = (
    "Cover image URL",
    "Back image URL",
    "Scanner ID",
)

BLANK_FILLS = {
    None,
    "00000000",
    "FF000000",
    "000000",
    "FFFFFFFF",
    "FFFFFF",
}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return " ".join(text.split())


def _fill_rgb(cell: Cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.patternType not in {"solid", "mediumGray", "darkGray", "lightGray"}:
        return None
    color = fill.fgColor
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        return str(color.rgb).upper()
    if color.type == "theme":
        return f"theme:{color.theme}:{color.tint}"
    return None


def is_colored_header(cell: Cell) -> bool:
    rgb = _fill_rgb(cell)
    if rgb is None:
        return False
    if rgb in BLANK_FILLS:
        return False
    if rgb.startswith("THEME:") and ":0.0" in rgb:
        return False
    return True


class CatalogWorkbook:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.path}")
        self.workbook: Workbook = load_workbook(self.path)
        self.sheet: Worksheet = self.workbook.active
        self._ensure_required_columns()
        self.all_columns: list[dict[str, Any]] = self._detect_all_columns()
        self.columns: list[dict[str, Any]] = [col for col in self.all_columns if col["colored"]]

    def _ensure_required_columns(self) -> None:
        by_name = {str(cell.value or "").strip(): cell for cell in self.sheet[1] if cell.value}
        template = self.sheet["D1"]
        column = self.sheet.max_column
        for header in REQUIRED_COLORED_HEADERS:
            cell = by_name.get(header)
            if cell is None:
                column += 1
                cell = self.sheet.cell(1, column, header)
                if template.has_style:
                    cell.font = copy(template.font)
                    cell.border = copy(template.border)
                    cell.alignment = copy(template.alignment)
                    cell.number_format = template.number_format
                    cell.protection = copy(template.protection)
                self.sheet.column_dimensions[get_column_letter(column)].width = 28
            cell.fill = ORANGE_FILL

    def _detect_all_columns(self) -> list[dict[str, Any]]:
        columns: list[dict[str, Any]] = []
        for cell in self.sheet[1]:
            header = cell.value
            if header is None:
                continue
            name = str(header).strip()
            columns.append(
                {
                    "index": cell.column,
                    "letter": get_column_letter(cell.column),
                    "header": name,
                    "field": HEADER_TO_FIELD.get(_normalize_header(name)),
                    "colored": is_colored_header(cell),
                }
            )
        return columns

    @property
    def colored_headers(self) -> list[str]:
        return [col["header"] for col in self.columns]

    def existing_keys(self) -> set[str]:
        isbn_col = next((c["index"] for c in self.columns if c["field"] == "isbn"), None)
        title_he_col = next((c["index"] for c in self.columns if c["field"] == "title_he"), None)
        title_en_col = next((c["index"] for c in self.columns if c["field"] == "title_en"), None)
        scanner_col = next((c["index"] for c in self.all_columns if c["field"] == "scanner_id"), None)
        keys: set[str] = set()
        for row in range(2, self.sheet.max_row + 1):
            if scanner_col:
                scanner_id = str(self.sheet.cell(row, scanner_col).value or "").strip()
                if scanner_id:
                    keys.add(f"scanner:{scanner_id}")
            if isbn_col:
                isbn = self.sheet.cell(row, isbn_col).value
                if isbn:
                    keys.add(f"isbn:{str(isbn).strip()}")
            title = ""
            if title_he_col:
                title = str(self.sheet.cell(row, title_he_col).value or "").strip()
            if not title and title_en_col:
                title = str(self.sheet.cell(row, title_en_col).value or "").strip()
            if title:
                keys.add(f"title:{title.casefold()}")
        return keys

    def existing_scanner_ids(self) -> set[str]:
        return {key[8:] for key in self.existing_keys() if key.startswith("scanner:")}

    def next_empty_row(self) -> int:
        for row in range(2, self.sheet.max_row + 2):
            values = [
                self.sheet.cell(row, col["index"]).value
                for col in self.columns
            ]
            if all(value in (None, "") for value in values):
                return row
        return self.sheet.max_row + 1

    def append_books(self, books: list[dict[str, Any]]) -> tuple[int, int, list[str], list[str]]:
        """Write selected books into colored columns only. Returns (written, skipped, written ids, skipped ids)."""
        existing = self.existing_keys()
        row = self.next_empty_row()
        written = 0
        skipped = 0
        written_ids: list[str] = []
        skipped_ids: list[str] = []
        for book in books:
            scanner_id = str(book.get("scanner_id") or "").strip()
            isbn = str(book.get("isbn") or "").strip()
            title = str(book.get("title_he") or book.get("title_en") or "").strip()
            if scanner_id and f"scanner:{scanner_id}" in existing:
                skipped += 1
                skipped_ids.append(scanner_id)
                continue
            if isbn and f"isbn:{isbn}" in existing:
                skipped += 1
                if scanner_id:
                    skipped_ids.append(scanner_id)
                continue
            if title and f"title:{title.casefold()}" in existing:
                skipped += 1
                if scanner_id:
                    skipped_ids.append(scanner_id)
                continue
            for col in self.columns:
                field = col["field"]
                if not field:
                    continue
                value = book.get(field)
                if value in (None, ""):
                    continue
                self.sheet.cell(row, col["index"], value)
            if scanner_id:
                existing.add(f"scanner:{scanner_id}")
                written_ids.append(scanner_id)
            if isbn:
                existing.add(f"isbn:{isbn}")
            if title:
                existing.add(f"title:{title.casefold()}")
            row += 1
            written += 1
        return written, skipped, written_ids, skipped_ids

    def remove_scanner_ids(self, scanner_ids: set[str] | list[str]) -> int:
        wanted = {str(item).strip() for item in scanner_ids if str(item).strip()}
        if not wanted:
            return 0
        scanner_col = next((c["index"] for c in self.all_columns if c["field"] == "scanner_id"), None)
        if not scanner_col:
            return 0
        rows = [
            row
            for row in range(2, self.sheet.max_row + 1)
            if str(self.sheet.cell(row, scanner_col).value or "").strip() in wanted
        ]
        for row in reversed(rows):
            self.sheet.delete_rows(row, 1)
        return len(rows)

    def save(self) -> Path:
        try:
            self.workbook.save(self.path)
            return self.path
        except PermissionError:
            fallback = self.path.with_name(f"{self.path.stem} - filled{self.path.suffix}")
            self.workbook.save(fallback)
            return fallback


_UNSAFE_FILENAME = re.compile(r'[<>:"/\\|?*]')


def list_excel_filename(title: str, list_id: str = "") -> str:
    cleaned = _UNSAFE_FILENAME.sub("_", (title or "New").strip())
    cleaned = " ".join(cleaned.split()).strip(" .") or "New"
    cleaned = cleaned[:80]
    suffix = str(list_id or "").strip()
    if suffix:
        return f"SISU - {cleaned[:68]} [{suffix[:12]}].xlsx"
    return f"SISU - {cleaned}.xlsx"


def ensure_list_workbook(path: str | Path, schema: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists():
        return target
    source = Path(schema)
    if not source.exists():
        raise FileNotFoundError(f"Excel schema was not found: {source}")
    shutil.copy2(source, target)
    return target
